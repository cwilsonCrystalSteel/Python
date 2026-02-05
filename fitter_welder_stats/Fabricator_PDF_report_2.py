# -*- coding: utf-8 -*-
"""
Created on Thu Feb  5 11:22:14 2026

@author: Netadmin
"""



import os
import pandas as pd
from pathlib import Path
import numpy as np 
import datetime
from openpyxl import Workbook, load_workbook
#%%

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, PageBreak, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import NextPageTemplate
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame


styles = getSampleStyleSheet()
title_style = ParagraphStyle(
    'TitleStyle',
    parent=styles['Title'],
    fontSize=16,
    alignment=1,  # Center align
    spaceAfter=12
)

centered_note_style = ParagraphStyle(
    name='CenteredNote',
    parent=styles['Normal'],
    alignment=TA_CENTER,
    fontSize=10,
    fontName='Helvetica-Oblique',  # Use the italic version of Helvetica
)


class pdf_report_fabricator():
    def __init__(self, state, csv_file=None, output_pdf=None, aggregate_data=None, past_agg_data=None, output_xlsx=None):
        self.state = state
        self.aggregate_data = aggregate_data
        if csv_file is None and aggregate_data is not None:
            self.csv_file = self.aggregate_data['filepath']
            self.get_bad_dfs()
            
        elif csv_file is not None:
            self.csv_file = csv_file
            
        elif csv_file is None and aggregate_data is None:
            raise Exception("Did not pass either 'csv_file' or 'aggregate_data'")
            
        self.past_agg_data = past_agg_data.copy()
         
        
            
            
        self.extrapolate_month_year()
        self.main_df = self.load_df(self.csv_file)
        #if we have past aggregate data files to look at 
        if not self.past_agg_data is None :
            # go thru each key in the past agg data and convert to df
            for k in self.past_agg_data.keys():
                # convert filepaths to nice dataframes 
                self.past_agg_data[k] = self.load_df(self.past_agg_data[k])
        
            # if we don't already have the current months df in the past agg dict then add id 
            if self.past_agg_data.get(0) is None:
                self.past_agg_data[0] = self.main_df.copy()
            
        
        self.output_pdf = output_pdf
        self.output_xlsx = output_xlsx
        self._excel_workbook = None
        
        if not os.path.dirname(output_pdf):
            os.makedirs(os.path.dirname(output_pdf))
            
        if not os.path.dirname(output_xlsx):
            os.makedirs(os.path.dirname(output_xlsx))
        
        # This will store the current values to display in the header
        self.header_context = {
            'state': '',
            'classification': '',
            'month': '',
            'year': ''
        }
        
        self.elements = []
        
        
        
        
        
    def build_report(self):
        self.init_document()
    
        # Title page (no header)
        self.elements.append(NextPageTemplate('NoHeader'))
        self.add_TitlePage()
        
        
        # setup the headers 
        # you kinda have to add the header template after the plot, to set it on that page?
        self.add_template("Fitter", "Fitter")
        self.add_template("Welder", "Welder")
        self.add_template("Combination", "Combination")
        self.add_template("AllEmployees", "Fitters & Welders")
        self.add_landscape_template('Landscape','Combination')
        
        
        
        # Page 1 - Different tpyes of hours for all employees 
        self.elements.append(NextPageTemplate("AllEmployees"))
        # self.do_pagebreak() # this is handled in the looping for the below graphic
        self.add_AllHourTypeComparison_n_per_page()
        # self.add_AllHourTypeComparison()
        
        classification = 'Combination'
        self.do_pagebreak()
        self.add_MonthOverMonth_Hours('Total Hours', classification)
        
        # self.do_pagebreak()
        self.add_MonthOverMonth_Hours('Missed Hours', classification)
        
        
        
        self.elements.append(NextPageTemplate("Landscape"))
        
        # Page 1 landscape
        self.do_pagebreak()
        self.add_state_table2()
        
        
        self.elements.append(NextPageTemplate(classification))
       
        # # page 2a
        # self.do_pagebreak()
        # self.add_MonthOverMonth_Hours('Total Hours', classification)
        
        # # page 2b
        # self.add_MonthOverMonth_Hours('Missed Hours', classification)
        
    
        
        # Page 3 defects
        self.do_pagebreak()
        # self.add_DefectsPlot(classification)
        self.add_DefectsPlot_both()
        
        # Page 4 Tons per piece 
        self.do_pagebreak()
        self.add_AverageTonsPerPiece(classification)
        
        # Page 5 efficienies
        self.do_pagebreak()
        self.add_DirectAndTotalEfficiency(classification)
        
        # Page 6 Earned hours Total/Direct
        self.do_pagebreak()
        self.add_EarnedAndTotalHours(classification)
        
       
        # Page 7 tonnage
        self.do_pagebreak()
        self.add_TonsByEmployee(classification)
        
        # # show specific class based 
        # for classification in ['Fitter','Welder']:
        #     # table for that classifcation
        #     self.do_pagebreak()
        #     self.add_state_table(classification)
            
        #     # earned hours graphic for that classification
        #     self.do_pagebreak()
        #     self.add_EarnedAndTotalHours(classification)
        
            
        # bad entries by fit/weld
        for classification in ['Fitter','Welder']:
            # self.elements.append(NextPageTemplate(classification))
            # self.do_pagebreak()
            # how bad the bad entries effect them
            self.do_pagebreak()
            self.add_badEntries(classification)
            
                
        self.build_document()
        self.save_excel_if_needed()

    
        
        
    def get_bad_dfs(self):
        self.bad_dfs = {}
        classes = ['fitters','welders']
        for classification in classes:
            df_wrong_state = self.aggregate_data[f'wrong_state_{classification}'][self.state]
            df_bad_id = self.aggregate_data[f'fix_{classification}'][self.state]
            df_didnt_work_that_month = self.aggregate_data[f'employee_didnt_work_{classification}'][self.state]
            
            if not df_wrong_state is None:
                df_wrong_state['reason'] = 'Employee Works at Different Shop'
            if not df_bad_id is None:
                df_bad_id['reason'] = 'Invalid Employee ID'
            if not df_didnt_work_that_month is None:
                df_didnt_work_that_month['reason'] = 'Employee Did Not Work in The Month'
            
            
            dfs = [df_wrong_state, df_bad_id, df_didnt_work_that_month]
            dfs = [i for i in dfs if not i is None]
            dfs_together = pd.concat(dfs, axis=0)
            
            
            
            if classification == 'fitters':
                self.bad_dfs['Fitter']  = dfs_together
                
            if classification == 'welders':
                self.bad_dfs['Welder'] = dfs_together
        
        
    def load_df(self, filepath):
        
        self.display_cols = ['Name','Earned Hours','Tons','Defect Quantity','Tons per Piece',
                        'Direct Hours','Total Hours','Missed Hours']

        # load the file
        main_df = pd.read_csv(filepath, index_col=0)
        # this is a place holder until I figure out how to alert on when employees are terminated but being credited with work /
        # did not work any during that month but are credited
        main_df = main_df[~main_df['direct'].isna()]

        main_df = main_df.sort_values('Earned Hours', ascending=False)

        main_df = main_df.rename(columns={'direct':'Direct Hours',
                                          'indirect':'Indirect Hours',
                                          'total':'Total Hours',
                                          'missed':'Missed Hours',
                                          'notcounted':'Other Hours',
                                          'Tonnage':'Tons',
                                          'Tonnage per Piece':'Tons per Piece'})

        return main_df
        
        
        
        
    def extrapolate_month_year(self):
        str_file = os.path.basename(self.csv_file)
        split_name = str_file.split('_')
        if 'prod_' in str_file:
            split_name = split_name[2].split('-')
        else:
            split_name = split_name[1].split('-')
        self.month_name = split_name[0]
        self.year = split_name[1]
        # date = datetime.datetime.strptime(month_name, '%B')
        # self.month_name = datetime.date(year=2000, month=int(month_number), day=1).strftime('%B')
        # self.year = split_name.split('_')[1]
        
        

        

    def init_document(self):
        # how to format the type of the output_pdf b/c we cant pass it a Path()
        output_pdf = str(self.output_pdf) if isinstance(self.output_pdf, Path) else self.output_pdf
        # generaet doc
        self.doc = BaseDocTemplate(output_pdf, pagesize=A4)
        # init the no header page template
        frame = Frame(36, 36, A4[0] - 72, A4[1] - 72, id='normal')
        self.doc.addPageTemplates([
            PageTemplate(id='NoHeader', frames=frame)
        ])   
        


        # --- Create PDF ---
        # doc = SimpleDocTemplate(output_pdf, pagesize=A4)
        
        
    def build_document(self):
        self.doc.build(self.elements)
        print(f'File saved to: {self.output_pdf}')
        
    

        
    def make_header_func(self, state, classification, month, year):
        def header(canvas_obj, doc):
            canvas_obj.saveState()
            canvas_obj.setFont("Helvetica", 10)
    
            left = f"{state} {classification}".strip()
            right = f"{month} {year}".strip()
    
            canvas_obj.drawString(40, A4[1] - 40, left)
            text_width = canvas_obj.stringWidth(right, "Helvetica", 10)
            canvas_obj.drawString(A4[0] - 40 - text_width, A4[1] - 40, right)
    
            canvas_obj.restoreState()
        return header

    def add_template(self, name, classification):
        frame = Frame(36, 36, A4[0] - 72, A4[1] - 72, id='normal')
        header_func = self.make_header_func(
            self.state,
            classification,
            self.month_name,
            self.year
        )
        self.doc.addPageTemplates([
            PageTemplate(id=name, frames=frame, onPage=header_func)
        ])


    
    def add_landscape_template(self, name, classification):
        pagesize = landscape(A4)
    
        frame = Frame(
            36, 36,
            pagesize[0] - 72,
            pagesize[1] - 72,
            id='landscape'
        )
    
        header_func = self.make_header_func(
            self.state,
            classification,
            self.month_name,
            self.year
        )
    
        self.doc.addPageTemplates([
            PageTemplate(
                id=name,
                frames=frame,
                pagesize=pagesize,
                onPage=header_func
            )
        ])
        
        
    
    def apply_header_template(self, classification):
        self.header_context['state'] = self.state
        self.header_context['classification'] = classification
        self.header_context['month'] = self.month_name
        self.header_context['year'] = self.year
    
    
    def do_pagebreak(self):
        self.elements.append(PageBreak())
        
    def add_TitlePage(self):
        print('Building add_TitlePage...')

        today = datetime.datetime.now()
        formatted_date = today.strftime('%B %d, %Y')  # e.g., April 21, 2025

        # --- Title and Subtitle Styles ---
        title_style = ParagraphStyle(
            name='Title',
            fontSize=32,
            alignment=1,  # Center
            spaceAfter=20
        )
        
        subtitle_style = ParagraphStyle(
            name='Subtitle',
            fontSize=20,
            alignment=1,  # Center
            spaceAfter=40
        )

        footer_style = ParagraphStyle(
            name='Footer',
            fontSize=10,
            alignment=1  # Center
        )

        # --- Add vertical space to center content ---
        self.elements.append(Spacer(1, 180))  # Push content toward vertical center

        # --- Main Title ---
        title = Paragraph(f"Fabricator Stats for {self.state}", title_style)
        self.elements.append(title)

        # --- Subtitle ---
        subtitle = Paragraph(f"{self.month_name} {self.year}", subtitle_style)
        self.elements.append(subtitle)

        # --- Footer (push to bottom of page) ---
        self.elements.append(Spacer(1, 300))  # Adjust as needed based on your page size
        footer = Paragraph(f"Report Generated on {formatted_date}", footer_style)
        self.elements.append(footer)
        self.elements.append(Spacer(1, 12)) 
        quesetions = Paragraph("Questions? Comments? cwilson@crystalsteel.net", footer_style)
        self.elements.append(quesetions)

        # --- Page break to end title page ---
        # self.do_pagebreak()
        
        
        
    def add_state_table(self, classification):
        print('Building add_state_table...')

        state_df = self.main_df[(self.main_df['Location'] == self.state) & 
                                (self.main_df['Classification'] == classification)].copy()
        # only select columns
        state_df = state_df[self.display_cols]
        # only show people who have earned something
        state_df = state_df[state_df['Earned Hours'] > 0]
        state_df['DL Efficiency']  = state_df['Earned Hours'] / state_df['Direct Hours']
        state_df['TTL Efficiency'] = state_df['Earned Hours'] / state_df['Total Hours']
        
        state_df['DL Efficiency'] = state_df['DL Efficiency'].replace(np.inf, 0)
        state_df['TTL Efficiency'] = state_df['TTL Efficiency'].replace(np.inf, 0)
        
        state_df_display = state_df.copy()
        rounding_display_dict = {'Earned Hours':'.0f',
                                 'Tons':'.1f',
                                 'Defect Quantity':'.2f',
                                 'Tons per Piece':'.3f',
                                 'Direct Hours':'.0f',
                                 'Total Hours':'.0f',
                                 'Missed Hours':'.0f',
                                 'DL Efficiency':'.0%',
                                 'TTL Efficiency':'.0%'}
        
        
        for col in rounding_display_dict.keys():
            fmt = rounding_display_dict[col]
            state_df_display[col] = state_df_display[col].apply(lambda x: f"{x:{fmt}}")
            
        state_df_display['Name'] = state_df_display['Name'].apply(lambda x: x if len(x) <= 20 else x[:20] + '...')
        
        
        col_headers = state_df_display.columns.tolist()
        col_headers = [Paragraph(i.replace(' ','<br/>')) for i in col_headers]
        

        
        # Convert DataFrame to list-of-lists
        data = [col_headers] + state_df_display.values.tolist()
        
        colWidths = [130,47,35,50,40,40,40,44,58,58]
        # Create Table
        pdf_table = Table(data, colWidths=colWidths)
        pdf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        # --- Add title above table ---
        title_text = f"Table Summary of {classification} Stats"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        # add the table in
        self.elements.append(pdf_table)
        
        note_text = (
        f"The values of Earned Hours, Tons, Defects, are only in relation to when the employee is "
        f"credited as being the {classification} for pieces. Again, these are pieces that the employee "
        f"has earned from fabricating as a {classification.upper()}. "
        f"As of August 2025: Earned Hours will correspond to the hours earned under the {classification} "
        "task, instead of the overal EVA hours for a piece!"
        )
        table_note = Paragraph(note_text, centered_note_style)
        self.elements.append(Spacer(1, 12))
        self.elements.append(table_note)
        
        
        self._write_excel_sheet(f"{classification} Summary Table", state_df)
        
            
    def add_state_table2(self):
        print('Building add_state_table2...')
    
        state_dfs = {}
        for classification in ['Combination','Fitter','Welder']:
            state_dfs[classification] = self.main_df[(self.main_df['Location'] == self.state) & 
                                    (self.main_df['Classification'] == classification)].copy()
    
        state_df = state_dfs['Combination']
        
        for classifcation in ['Fitter','Welder']:
            classification_df = state_dfs[classifcation]
            state_df = pd.merge(left=state_df,
                     right=classification_df,
                     left_on='Name',
                     right_on='Name',
                     suffixes=('',f'_{classifcation}'),
                     how='left')
        
        
        
        overall_cols = ['Name','Direct Hours','Total Hours','Missed Hours']
        all_classification_cols = ['Earned Hours','Tons','Tons per Piece']
        cols_to_use = [i for i in overall_cols] + [i for i in state_df.columns for j in all_classification_cols if i.startswith(j)]
        cols_to_use = [i for i in cols_to_use if 'per defect' not in i.lower()]
        
        # only select columns
        state_df = state_df[cols_to_use]
        
        rename_cols = [i for i in state_df.columns if '_' in i]
        rename_cols_dict = {i:i.split('_')[1] + ' ' + i.split('_')[0] for i in rename_cols}
        state_df = state_df.rename(columns=rename_cols_dict)
        
        # only show people who have earned something
        state_df = state_df[state_df['Earned Hours'] > 0]
        state_df['DL Efficiency']  = state_df['Earned Hours'] / state_df['Direct Hours']
        state_df['TTL Efficiency'] = state_df['Earned Hours'] / state_df['Total Hours']
        
        state_df['DL Efficiency'] = state_df['DL Efficiency'].replace(np.inf, 0)
        state_df['TTL Efficiency'] = state_df['TTL Efficiency'].replace(np.inf, 0)
        
        
        to_move = state_df.pop('DL Efficiency')
        state_df.insert(4, 'DL Efficiency', to_move)
        
        to_move = state_df.pop('TTL Efficiency')
        state_df.insert(5, 'TTL Efficiency', to_move)
        
        state_df = state_df.loc[:, ~state_df.columns.duplicated()].copy()

        
        state_df_display = state_df.copy()
        rounding_display_dict = {'Earned Hours':'.0f',
                                 'Tons':'.1f',
                                 'Defect Quantity':'.2f',
                                 'Tons per Piece':'.3f',
                                 'Direct Hours':'.0f',
                                 'Total Hours':'.0f',
                                 'Missed Hours':'.0f',
                                 'DL Efficiency':'.0%',
                                 'TTL Efficiency':'.0%'}
        
        for col in state_df_display.columns:
            fmt = None
            for fmt_col in rounding_display_dict.keys():
                if fmt_col in col:
                    fmt = rounding_display_dict[fmt_col]
                    break
            if fmt is not None:
                state_df_display[col] = state_df_display[col].apply(lambda x: f"{x:{fmt}}")
                
        # for col in rounding_display_dict.keys():
        #     fmt = rounding_display_dict[col]
        #     state_df_display[col] = state_df_display[col].apply(lambda x: f"{x:{fmt}}")
            
        state_df_display['Name'] = state_df_display['Name'].apply(lambda x: x if len(x) <= 20 else x[:20] + '...')
        
        
        col_headers = state_df_display.columns.tolist()
        col_headers = [Paragraph(i.replace(' ','<br/>')) for i in col_headers]
        
        eff_cols = ['DL Efficiency', 'TTL Efficiency']
        eff_col_idxs = [
            state_df_display.columns.get_loc(c)
            for c in eff_cols
            if c in state_df_display.columns
        ]

        
        # Convert DataFrame to list-of-lists
        data = [col_headers] + state_df_display.values.tolist()
        
        colWidths = [130,47,35,50,40,40,40,44,58,58,55,55,55,55,55]
        # Create Table
        pdf_table = Table(data, colWidths=colWidths)
        # pdf_table.setStyle(TableStyle([
        #     ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        #     ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        #     ('GRID', (0, 0), (-1, -1), 1, colors.black),
        #     ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #     ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        # ]))
        
        
                
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]
        
        # Shade efficiency columns (data rows only)
        for idx in eff_col_idxs:
            table_style.append(
                ('BACKGROUND', (idx, 1), (idx, -1), colors.mintcream)
            )
        
        pdf_table.setStyle(TableStyle(table_style))
        
        
        
        # --- Add title above table ---
        title_text = "Table Summary of Stats"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        # add the table in
        self.elements.append(pdf_table)
        
        # note_text = (
        # f"The values of Earned Hours, Tons, Defects, are only in relation to when the employee is "
        # f"credited as being the {classification} for pieces. Again, these are pieces that the employee "
        # f"has earned from fabricating as a {classification.upper()}. "
        # f"As of August 2025: Earned Hours will correspond to the hours earned under the {classification} "
        # "task, instead of the overal EVA hours for a piece!"
        # )
        # table_note = Paragraph(note_text, centered_note_style)
        # self.elements.append(Spacer(1, 12))
        # self.elements.append(table_note)
        
        
        self._write_excel_sheet("Summary Table", state_df)        
        
        
    def add_AllHourTypeComparison_n_per_page(self):
        print('Building add_AllHourTypeComparison_n_per_page')
        from fitter_welder_stats.fitter_welder_stats_graphing import hours_comparison_by_employee
        
        df = self.main_df.copy()
        # i just copied these header names from: 
        # from fitter_welder_stats.fitter_welder_stats_graphing import hours_comparison_by_employee
        col0 = 'Name'
        col00 = 'Location'
        col1 = 'Total Hours'
        col2 = 'Direct Hours'
        col3 = 'Indirect Hours'
        col4 = 'Other Hours'
        col5 = 'Missed Hours'
        
        df = df[(df['Earned Hours'] > 0) | (df['Weight'] > 0) | (df['Quantity'] > 0)]
        
        # only keep cols we want
        df = df[[col0, col00, col1, col2, col3, col4, col5]]
        # just get one entry per worker
        df = df.drop_duplicates(keep='first', subset=['Name','Location'])
        
        # only get this state b/c we are doing this page-end stuff
        df = df[df['Location'] == self.state]
        # only get more than 0 hours worked
        df = df[df[col1] > 0]
        # sort now 
        df = df.sort_values(col1, ascending=False)
        
      
        
        total = df.shape[0]
        min_per_page = 18
        min_per_page = min_per_page if min_per_page < total else total 
        max_per_page = 32
        
        possible_pages = []
        
        for pages in range(1, total + 1):
            per_page = total / pages
            if min_per_page <= per_page <= max_per_page:
                possible_pages.append(pages)
        
        # pick the page count closest to the midpoint
        target = (min_per_page + max_per_page) / 2
        num_pages = min(
            possible_pages,
            key=lambda p: abs((total / p) - target)
        )
        
        base = total // num_pages
        remainder = total % num_pages
        
        page_sizes = [
            base + (1 if i < remainder else 0)
            for i in range(num_pages)
        ]
        
        
        start = 0
        
        for size in page_sizes:
            end = start + size
            print(f'Building add_AllHourTypeComparison_n_per_page for #{start}-{end}')
            fake_classifcation_for_graphic_saving = f'{start}-{end}'
            graphic, _ = hours_comparison_by_employee(main_df = df, 
                                                            state=self.state, 
                                                            classification=fake_classifcation_for_graphic_saving, 
                                                            topN=(start,end), 
                                                            SAVEFILES=True)
            if not os.path.exists(graphic):
                raise Exception('Could not find {graphic}')
                
            
            self.do_pagebreak()
            title_text = f"Types of Hours Worked (Rank #{start+1}-{end})"
            title = Paragraph(title_text, title_style)
            self.elements.append(title)
            img = Image(graphic, width=520, height=640)
            self.elements.append(img)
        
            
            # move the index range up 
            start = end
            
            
        # still do this so that we can add to the excel file
        _, xlsx_df = hours_comparison_by_employee(self.main_df, self.state, SAVEFILES=True)
        self._write_excel_sheet("Types of Hours Worked", xlsx_df)
        
        
    def add_AllHourTypeComparison(self):
        print('Building add_AllHourTypeComparison...')
        from fitter_welder_stats.fitter_welder_stats_graphing import hours_comparison_by_employee
        graphic, xlsx_df = hours_comparison_by_employee(self.main_df, self.state, SAVEFILES=True)
        if not os.path.exists(graphic):
            raise Exception('Could not find {graphic}')


        title_text = "Classification of Hours Worked"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        
        img = Image(graphic, width=520, height=640)
        
        self.elements.append(img)
        
        # note_text = (
        # f"* Employees with more than {min_qty} pieces credited as a {classification}."
        # )
        # table_note = Paragraph(note_text, centered_note_style)
        # self.elements.append(Spacer(1, 12))
        # self.elements.append(table_note)
        self._write_excel_sheet("Hours Worked by Type", xlsx_df)
        
    def add_MonthOverMonth_Hours(self, hours_type, classification):
        print(f'Building add_MonthOverMonth_Hours {hours_type} {classification}')
        # HOURS_TYPE_VALID = ['Total Hours','Direct Hours','Missed Hours']
        from fitter_welder_stats.fitter_welder_stats_graphing import mom_hours_worked_by_shop
        graphic = mom_hours_worked_by_shop(dict_of_dfs = self.past_agg_data, 
                                           state = self.state, 
                                           hours_type = hours_type, 
                                           classification = classification,
                                           month_0_name = self.month_name, 
                                           month_0_year = self.year,
                                           SAVEFILES=True)
        
        title_text = f'Average {hours_type} Worked by Month'
        if not classification is None:
            title_text += f" for {classification}"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        img = Image(graphic, width=520, height = 325)
        
        self.elements.append(img)
        
        # self._write_excel_sheet(f"{classification} Monthly Hours", state_df)
        
        
    def add_DefectsPlot(self, classification):
        print(f'Building add_DefectsPlot {classification}')
        from fitter_welder_stats.fitter_welder_stats_graphing import defects_by_employee
        graphic = defects_by_employee(self.main_df, self.state, classification, SAVEFILES=True)
        if not os.path.exists(graphic):
            raise Exception('Could not find {graphic}')
            
        title_text = f"{classification} Type Defects"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        img = Image(graphic, width=487.5, height=650)
        
        self.elements.append(img)
        
    def add_DefectsPlot_both(self, ):
        print('Building add_DefectsPlot_both')
        from fitter_welder_stats.fitter_welder_stats_graphing import defects_by_employee_both_classifications
        graphic = defects_by_employee_both_classifications(self.main_df, self.state, SAVEFILES=True)
        if not os.path.exists(graphic):
            raise Exception('Could not find {graphic}')
            
        title_text = " Defects"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        img = Image(graphic, width=487.5, height=650)
        
        self.elements.append(img)        
        
        
    def add_AverageTonsPerPiece(self, classification):
        print(f'Building add_AverageTonsPerPiece {classification}')
        from fitter_welder_stats.fitter_welder_stats_graphing import tonnage_per_piece_by_employee
        min_qty = 5
        graphic = tonnage_per_piece_by_employee(self.main_df, self.state, classification, topN=25, min_qty=min_qty, SAVEFILES=True)    
        if not os.path.exists(graphic):
            raise Exception('Could not find {graphic} ')
                    
            
        title_text = f"{classification} Tons Per Piece Completed"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        
        img = Image(graphic, width=520, height=650)
        
        self.elements.append(img)
        
        note_text = (
        f"* Employees with more than {min_qty} pieces credited as a {classification}."
        )
        table_note = Paragraph(note_text, centered_note_style)
        self.elements.append(Spacer(1, 12))
        self.elements.append(table_note)
        
        
        
        
    def add_DirectAndTotalEfficiency(self, classification):
        print(f'Building add_DirectAndTotalEfficiency {classification}...')
        min_hours = 10
        from fitter_welder_stats.fitter_welder_stats_graphing import total_direct_labor_efficiency
        graphic = total_direct_labor_efficiency(self.main_df, self.state, classification, min_hours=min_hours, SAVEFILES=True)   
        if not os.path.exists(graphic):
            raise Exception('Could not find {graphic}')
            
        title_text = f"{classification} Efficiency (Based on Direct & Total Hours)"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        img = Image(graphic, width=520, height=650)
        
        self.elements.append(img)
        
        note_text = (
        f"* Employees with more than {min_hours} Direct Hours worked."
        )
        table_note = Paragraph(note_text, centered_note_style)
        self.elements.append(Spacer(1, 12))
        self.elements.append(table_note)
        
        
    def add_EarnedAndTotalHours(self, classification):
        print(f'Building add_EarnedAndTotalHours {classification}...')
        min_hours = 10
        from fitter_welder_stats.fitter_welder_stats_graphing import earned_hours_by_employee
        graphic = earned_hours_by_employee(self.main_df, self.state, classification, min_hours=min_hours, SAVEFILES=True)
        if not os.path.exists(graphic):
            raise Exception('Could not find {graphic}')
            
        title_text = f"{classification} Earned Hours & Hours Worked"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        img = Image(graphic, width=480, height=600)
        
        self.elements.append(img)
        
        note_text = (
        f"* Employees with more than {min_hours} Earned Hours credited as a {classification}."
        )
        table_note = Paragraph(note_text, centered_note_style)
        self.elements.append(Spacer(1, 12))
        self.elements.append(table_note)
        
        
    def add_TonsByEmployee(self, classification):
        print(f'Building add_TonsByEmployee {classification}...')
        min_tons = 3
        from fitter_welder_stats.fitter_welder_stats_graphing import weight_by_employee
        graphic = weight_by_employee(self.main_df, self.state, classification, SAVEFILES=True)
        if not os.path.exists(graphic):
            raise Exception('Could not find {graphic}')
            
        title_text = f"Tons Completed as {classification}"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        img = Image(graphic, width=480, height=600)
        
        self.elements.append(img)
        
        note_text = (
        f"* Employees with more than {min_tons} Tons credited as a {classification}."
        )
        table_note = Paragraph(note_text, centered_note_style)
        self.elements.append(Spacer(1, 12))
        self.elements.append(table_note)   
        
        
    def add_badEntries(self, classification):
        print(f'Building add_badEntries {classification}...')

        from fitter_welder_stats.fitter_welder_stats_graphing import bad_EmployeeCredit, bad_entriesPieChartAndTable
        
        graphic1 = bad_EmployeeCredit(self.bad_dfs[classification], self.state, classification, SAVEFILES=True)
        graphic2 = bad_entriesPieChartAndTable(self.main_df, self.bad_dfs[classification], self.state, classification, SAVEFILES=True)
        if not os.path.exists(graphic1['filepath']):
            raise Exception('Could not find {graphic1}')
            
        if not os.path.exists(graphic2['filepath']):
            raise Exception('Could not find {graphic2}')
            
        title_text = f"Effect of Bad Entries into Fablisting - {classification}"
        title = Paragraph(title_text, title_style)
        self.elements.append(title)
        
        img = Image(graphic1['filepath'], width=470, height=300)
        
        self.elements.append(img)
        self.elements.append(Spacer(1, 24))  # 12 points = approx 1 line
        
        img = Image(graphic2['filepath'], width=480, height=250)
        
        self.elements.append(img)
        
        self.elements.append(Spacer(1, 12))
        
        note_text = (
        "<b>* Invalid Employee ID</b>: No ID was entered, OR Employee ID did not match any ID in TimeClock."
        f"<br/><b>* Employee Did Not Work in The Month</b>: Employee ID returned an Employee who did not log hours in Timeclock for the month at hand."
        f"<br/><b>* Employee Works at Different Shop</b>: Employee ID returned an Employee who has a Productive Status indicating a different shop."
        )
        
        table_note = Paragraph(note_text, centered_note_style)
        self.elements.append(Spacer(1, 12))
        self.elements.append(table_note)           
        

    # ---------------------------
    # EXCEL HELPER METHODS
    # ---------------------------
    def _ensure_excel_workbook(self):
        """Creates or loads an Excel workbook only when needed."""
        if self.output_xlsx is None:
            return None

        if self._excel_workbook is None:
            excel_file = Path(self.output_xlsx)

            if excel_file.exists():
                self._excel_workbook = load_workbook(excel_file)
            else:
                self._excel_workbook = Workbook()
                del self._excel_workbook["Sheet"]  # remove default blank sheet

        return self._excel_workbook

    def _write_excel_sheet(self, sheet_name: str, df: pd.DataFrame):
        """Write a dataframe to a new sheet in the Excel file (if enabled)."""
        if self.output_xlsx is None:
            return  # Do nothing if Excel export is disabled

        wb = self._ensure_excel_workbook()

        # Avoid sheet-name conflicts
        if sheet_name in wb.sheetnames:
            # Auto-rename: e.g., "Summary", "Summary (2)"
            i = 2
            new_name = f"{sheet_name} ({i})"
            while new_name in wb.sheetnames:
                i += 1
                new_name = f"{sheet_name} ({i})"
            sheet_name = new_name[:30] # excel is funny 

        ws = wb.create_sheet(title=sheet_name)

        # Write dataframe to sheet
        ws.append(df.columns.tolist())
        for _, row in df.iterrows():
            ws.append(row.tolist())

    def save_excel_if_needed(self):
        """Finalize and write Excel workbook to disk."""
        if self.output_xlsx and self._excel_workbook:
            self._excel_workbook.save(self.output_xlsx)
            # self._excel_workbook = None
            print(f'File saved to: {self.output_xlsx}')
            

        
        

        

#%%



# x = pdf_report('MD', df_file, Path(r'C:/Users/Netadmin/Documents/FitterWelderStatsPDFReports/output2.pdf'))

# x.build_report()

