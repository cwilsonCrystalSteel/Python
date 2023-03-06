# -*- coding: utf-8 -*-
"""
Created on Tue May 25 14:15:13 2021

@author: CWilson
"""

from openpyxl import load_workbook
import pandas as pd
import sys
sys.path.append('c://users/cwilson/documents/python/attendance project')
from attendance_google_sheets_credentials_startup import init_google_sheet
import time


wb = load_workbook(filename = 'c://users/cwilson/downloads/CSF PRODUCTION WORKSHEET (20).xlsx')
sheet_names = wb.sheetnames
name = 'LOTS Log'
sheet_ranges = wb[name]
df = pd.DataFrame(sheet_ranges.values)
# df_headers = df.iloc[0,:].to_list()
formulas = df[9]
formulas = formulas.fillna("", inplace=True)
formulas = df[9]


sh = init_google_sheet("1HIpS0gbQo8q1Pwo9oQgRFUqCNdud_RXS3w815jX6zc4")
worksheet = sh.worksheet('LOTS Log')

cells_to_update = len(formulas)

for x in range(cells_to_update):
    # worksheet.update('J' + str(x + 1), formulas[x], value_input_option='USER_ENTERED')
    time.sleep(1.1)